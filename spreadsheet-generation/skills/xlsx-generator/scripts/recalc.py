#!/usr/bin/env python3
"""Recalculate formulas in an XLSX file and detect errors.

Uses LibreOffice headless to recalculate, then scans for formula errors.
Returns a JSON report with error locations and types.

Usage:
    python recalc.py <xlsx_file> [--output-report <report.json>]

Returns JSON:
    {
        "file": "model.xlsx",
        "error_count": 0,
        "error_cells": [],
        "recalc_success": true,
        "recalc_error": null
    }
"""

import json
import os
import sys


def find_formula_errors(xlsx_path):
    """Scan an XLSX file for formula error values.

    Args:
        xlsx_path: Path to the XLSX file.

    Returns:
        List of dicts with sheet, cell, formula, and error fields.
    """
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    errors = []
    error_values = {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A", "#NUM!", "#GETTING_DATA"}

    # Load with data_only=True to see calculated values
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in error_values:
                    errors.append({
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "error": cell.value,
                    })
    wb.close()

    # Also load without data_only to get the formulas for error cells
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
    for error in errors:
        ws = wb_formulas[error["sheet"]]
        cell = ws[error["cell"]]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
            error["formula"] = cell.value
        else:
            error["formula"] = None
    wb_formulas.close()

    return errors


def recalc_with_libreoffice(xlsx_path):
    """Recalculate formulas using LibreOffice headless.

    Args:
        xlsx_path: Path to the XLSX file.

    Returns:
        Tuple of (success: bool, error: str or None)
    """
    # Use the soffice utility from the resolved scripts
    soffice_script = os.path.join(
        os.getcwd(), ".reporting-resolved", "scripts", "office", "soffice.py"
    )

    if os.path.isfile(soffice_script):
        import subprocess
        result = subprocess.run(
            [sys.executable, soffice_script, "recalc", xlsx_path],
            capture_output=True, text=True, timeout=300,
        )
        if result.returncode == 0:
            return True, None
        else:
            return False, result.stderr.strip() or result.stdout.strip()

    # Direct LibreOffice fallback
    import shutil
    import subprocess

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return False, "LibreOffice not found. Install with: brew install --cask libreoffice"

    outdir = os.path.dirname(os.path.abspath(xlsx_path))
    cmd = [soffice, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", outdir, xlsx_path]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        return result.returncode == 0, result.stderr.strip() if result.returncode != 0 else None
    except subprocess.TimeoutExpired:
        return False, "Recalculation timed out after 5 minutes"


def recalc(xlsx_path, report_path=None):
    """Recalculate and validate an XLSX file.

    Args:
        xlsx_path: Path to the XLSX file.
        report_path: Optional path to write JSON report.

    Returns:
        Dict with recalc results and error report.
    """
    report = {
        "file": os.path.basename(xlsx_path),
        "error_count": 0,
        "error_cells": [],
        "recalc_success": False,
        "recalc_error": None,
    }

    # Step 1: Recalculate with LibreOffice
    success, error = recalc_with_libreoffice(xlsx_path)
    report["recalc_success"] = success
    report["recalc_error"] = error

    # Step 2: Scan for formula errors
    if success:
        errors = find_formula_errors(xlsx_path)
        report["error_count"] = len(errors)
        report["error_cells"] = errors
    elif not success and error and "not found" in error.lower():
        # LibreOffice not available — still scan for obvious errors
        report["recalc_error"] = f"{error} (formula values not recalculated, scanning for cached errors)"
        errors = find_formula_errors(xlsx_path)
        report["error_count"] = len(errors)
        report["error_cells"] = errors

    # Write report
    if report_path:
        with open(report_path, "w") as f:
            json.dump(report, f, indent=2)

    return report


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    report_path = None
    if "--output-report" in sys.argv:
        report_path = sys.argv[sys.argv.index("--output-report") + 1]

    if not os.path.isfile(xlsx_path):
        print(f"Error: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    result = recalc(xlsx_path, report_path)
    print(json.dumps(result, indent=2))
    sys.exit(0 if result["error_count"] == 0 else 1)
